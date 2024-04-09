import csv
import json
import sqlite3
from openpyxl import load_workbook

'''
    This function reads data from the hotel.csv file and returns a list of lists
    containing the data. Each row in the CSV file is read as a list and appended to the
    hotel list, which is then returned by the function.
'''
def get_hotel_data():
    with open('data/hotel.csv') as csvfile:
        hotel = []
        reader = csv.reader(csvfile)
        next(reader)
        for row in reader:
            if row and row[0] and row[1] and row[2] and row[3]:
                hotel.append([int(row[0]), int(row[1]), row[2], row[3]])
    return hotel

'''
    This function reads data from the client.csv file and returns a list of lists
    containing the data. Each row in the CSV file is read as a list and appended to the 
    client list, which is then returned by the function.
'''
def get_client_data():
    with open('data/client.csv') as csvfile:
        client = []
        reader = csv.reader(csvfile)
        next(reader)
        for row in reader:
            if row and row[0] and row[1] and row[2] and row[3] and row[4]:
                client.append([int(row[0]), row[1], row[2], int(row[3]), int(row[4])])
    return client

'''
    This function reads data from the data/roomunavailable.csv file and returns a list of 
    lists containing the data. Each row in the CSV file is read as a list and appended to 
    the room_used list, which is then returned by the function. The first row of the CSV 
    file is skipped because it is assumed to be a header row.
'''
def get_room_unavailable_data():
    with open('data/room_unavailable.csv') as csvfile:
        room_used = []
        reader = csv.reader(csvfile)
        next(reader)
        for row in reader:
            if row and row[0] and row[1] and row[2] and row[3]: # Ignore empty
                room_used.append([row[0], row[1], row[2], row[3]])
    return room_used

'''
    This function creates a connection with the Reservations.db file. A select query is executed 
    for each row in the Reservations table. A list of tuples is returned containing every row in the
    Reservations table.
'''
def get_reservations_data():
    con = sqlite3.connect("data/reservations.db")
    reservations = []
    cur = con.cursor()
    cur.execute("SELECT * FROM reserve")
    rows = cur.fetchall()
    for row in rows:
        if row and row[0] and row[1] and row[2] and row[3] and row[4] and row [5]:
            reservations.append([int(row[0]), int(row[1]), int(row[2]), float(row[3]), row[4], int(row[5])])
    cur.close()
    con.close()
    return reservations

'''
    This function creates a connection with the Rooms.db file. A select query is executed 
    for each row in the Rooms table. A list of tuples is returned containing every row in the
    Rooms table. 
'''
def get_rooms_data():
    con = sqlite3.connect("data/rooms.db")
    rooms = []
    cur = con.cursor()
    cur.execute("SELECT * FROM room")
    rows = cur.fetchall()
    for row in rows:
        if row and row[0] and row[1] and row[2] and row[3]:
            rooms.append([int(row[0]), int(row[1]), int(row[2]), float(row[3])])
    cur.close()
    con.close()
    return rooms

'''
    This function reads data from the data/employee.json file and returns a list of 
    lists containing the data. It loads the json file as a list of dictionaries, and checks 
    that each dictionary key has no empty values. Then it appends a list of the information 
    in each dictionary, to the employee list, and returns it.
'''
def get_employee_data():
    employee = []
    # Opens the employee json file to read it
    with open('data/employee.json', 'r') as f_json:
        data_json = json.load(f_json) # Creates a list with the contents of the json file
        # Iterates through the list
        for d in data_json:
            # Checks if any column is empty
            if d["employeeid"] != "" and d["hotelid"] != "" and d["firstname"] and d["lastname"] and d["age"] != "" and d["salary"] != "" and d["position"]:
                # Add each row in the employee list
                employee.append([int(d["employeeid"]), int(d["hotelid"]), d["firstname"], d["lastname"], int(d["age"]), float(d["salary"]), d["position"]])
    return employee

'''
    This function reads data from the data/roomdetails.json file and returns a list of 
    lists containing the data. It loads the json file as a list of dictionaries, and checks 
    that each dictionary key has no empty values. Then it appends a list with the information 
    in each dictionary, to the room_details list, and returns it.
'''
def get_room_details_data():
    room_details = []
    # Opens the roomdetails json file to read it
    with open('data/roomdetails.json', 'r') as f_json:
        data_json = json.load(f_json) # Creates a list with the contents of the json file
        # Iterates through the list
        for d in data_json:
            # Checks if any column is empty
            if d["detailid"] != "" and d["name"] and d["type"] and d["capacity"] != "" and d["handicap"] != "":
                # Add each row in the room_details list
                room_details.append([int(d["detailid"]), d["name"], d["type"], int(d["capacity"]), bool(int(d["handicap"]))])
    return room_details

'''
    This function, is designed to read data from an Excel file and return a list of rows,
    excluding those with 'None' values. The function then iterates over each row in the worksheet. The first row (header) 
    is skipped using the enumerate function. If the row does not contain 'None' or '', it is appended to the 'chain' list.
'''
def get_chain_data():
    chain = []
    data_file = 'data/chain.xlsx'
    wb = load_workbook(data_file) 
    ws = wb.active 
    
    for i, value in enumerate(ws.iter_rows(values_only = True), start=1): 
        if i == 1:  # Skip the first row (header)
            continue
        if None not in value and '' not in value:
            chain.append(value)
    
    return chain

'''
    This function, is designed to read data from an Excel file and return a list of rows,
    excluding those with 'None' values. The function then iterates over each row in the worksheet. The first row (header) 
    is skipped using the enumerate function. If the row does not contain 'None' or '', it is appended to the 'login' list.
'''
def get_login_data():
    
    login = []

    data_file = 'data/login.xlsx'
    wb = load_workbook(data_file)   
    ws = wb.active 
    
    for i, value in enumerate(ws.iter_rows(values_only = True ), start=1):
        if i == 1:
            continue
        if None not in value and '' not in value:
            login.append(value)
    
    return login
